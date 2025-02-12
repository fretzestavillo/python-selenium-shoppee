from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

import os
from dotenv import load_dotenv
from pyshadow.main import Shadow
import openpyxl
import smtplib
import time
import re

# Setup Chrome options
options = Options()
options.add_argument("start-maximized")
options.add_experimental_option(
    "excludeSwitches", ["enable-automation", "enable-logging"])
options.add_experimental_option("useAutomationExtension", False)
options.add_argument('--disable-blink-features=AutomationControlled')

# Set a realistic user-agent to avoid bot detection
options.add_argument(
    "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
)

# Start WebDriver
driver = webdriver.Chrome(service=Service(
    ChromeDriverManager().install()), options=options)

# Open Shopee login page
driver.get("https://shopee.ph/")

time.sleep(50)


# # # # # credential
# load_dotenv('.env')
# email: str = os.getenv('email')
# password: str = os.getenv('password')
# time.sleep(5)

# email_input = driver.find_element(By.CSS_SELECTOR, 'input[name="loginKey"]')
# email_input.send_keys(email)
# pass_input = driver.find_element(By.CSS_SELECTOR, 'input[name="password"]')
# pass_input.send_keys(password)

# time.sleep(10)

# login_form = driver.find_element(
#     By.CSS_SELECTOR, 'button[class="b5aVaf PVSuiZ Gqupku qz7ctP qxS7lQ Q4KP5g"]')
# login_form.click()
# time.sleep(10)

# verify_email = driver.find_element(
#     By.CSS_SELECTOR, 'button[class="SKUrUk"]')
# verify_email.click()

# time.sleep(50)


# actions = ActionChains(driver)
# actions.move_to_element(driver.find_element(
#     By.XPATH, "//div[text()='estafretz']"))
# actions.perform()


# time.sleep(2)

# purchase = driver.find_element(By.XPATH, "//span[text()='my purchase']")
# purchase.click()

# time.sleep(2)

# completed = driver.find_element(
#     By.XPATH, "//*[@id='main']/div/div[2]/div[1]/div[2]/div/div/div[2]/a[5]/span")
# completed.click()
# time.sleep(3)


# elements = driver.find_elements(By.CSS_SELECTOR, "span[class='x5GTyN']")
# costs = driver.find_elements(By.CSS_SELECTOR, "div[class='DeWpya']")


# # # Wait for the page to load and elements to be present
# wait = WebDriverWait(driver, 10)
# elements = wait.until(EC.presence_of_all_elements_located(
#     (By.CSS_SELECTOR, "span[class='x5GTyN']")))
# costs = wait.until(EC.presence_of_all_elements_located(
#     (By.CSS_SELECTOR, "div[class='DeWpya']")))

# # # Scroll to the bottom of the page
# last_height = driver.execute_script("return document.body.scrollHeight")
# while True:
#     driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
#     time.sleep(1)  # Adjust the delay as needed
#     new_height = driver.execute_script("return document.body.scrollHeight")
#     if new_height == last_height:
#         break
#     last_height = new_height

# # # Retrieve additional elements and costs at the bottom of the page
# elements = driver.find_elements(By.CSS_SELECTOR, "span[class='x5GTyN']")
# costs = driver.find_elements(By.CSS_SELECTOR, "div[class='DeWpya']")


# # # Process the retrieved elements and costs as needed
# total_cost = 0
# for element, cost in zip(elements, costs):
#     cost_value = int(cost.text.strip('₱').replace(',', ''))
#     total_cost += cost_value
#     print(f"Element: {element.text} - Cost: {cost.text}")

# print("Total Cost:", total_cost)


# time.sleep(3)

# workbook = openpyxl.Workbook()
# sheet = workbook.active

# # # Write the headers to the Excel file
# sheet.cell(row=1, column=1, value='Element')
# sheet.cell(row=1, column=2, value='Cost')

# # # Write the retrieved data to the Excel file and calculate the total cost
# total_cost = 0
# for i, (element, cost) in enumerate(zip(elements, costs), start=2):
#     cost_value = int(cost.text.strip('₱').replace(',', ''))
#     sheet.cell(row=i, column=1, value=element.text)
#     sheet.cell(row=i, column=2, value=cost_value)
#     total_cost += cost_value

# # # Write the total cost to the Excel file
# sheet.cell(row=i+1, column=1, value='Total Cost')
# sheet.cell(row=i+1, column=2, value=total_cost)


# # # Save the Excel file
# workbook.save('shopee.xlsx')


# # # credential
# load_dotenv('.env')
# email: str = os.getenv('gemail')
# password: str = os.getenv('gpass')
# receiver: str = os.getenv('greceiver1')

# # # Email configuration
# sender_email = email
# sender_password = password
# receiver_email = receiver
# subject = 'Data from Web Scraper'
# body = 'Please find the attached data file.'

# # # Create a multipart message and set the headers
# message = MIMEMultipart()
# message['From'] = sender_email
# message['To'] = receiver_email
# message['Subject'] = subject

# # # Attach the Excel file to the email
# attachment_path = 'shopee.xlsx'
# attachment = open(attachment_path, 'rb')
# part = MIMEBase('application', 'octet-stream')
# part.set_payload(attachment.read())
# encoders.encode_base64(part)
# part.add_header('Content-Disposition',
#                 f'attachment; filename="{attachment_path}"')
# message.attach(part)

# # # Convert the message to a string and send the email
# smtp_server = 'smtp.gmail.com'
# smtp_port = 587
# with smtplib.SMTP(smtp_server, smtp_port) as server:
#     server.starttls()
#     server.login(sender_email, sender_password)
#     server.send_message(message)

# # Close the attachment file
# attachment.close()
